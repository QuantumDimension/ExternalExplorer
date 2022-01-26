########################################################################################
### The Interactive Excel: External Explorer Script Part 2: Excel Creation           ###
### Creator: Glenn Nor                                                               ###
### Date of publication: January 2022                                                ###
### Version 2.13                                                                     ###
### This version is designed to be used with data from the digital forensic software ###
### AccessData Forensic ToolKit 7.5                                                  ###
########################################################################################


### IMPORT MODULES ########################################
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle, PatternFill
import os
import glob
import shutil
###########################################################

### Function to correct extension ###
def Extension(value):
    if value != None:
        new = "."+value
    if value == None:
        new = ".pdf"
    if value == "msg":
        new = ".pdf"
    return new
#####################################


### Fetch original item file list from FTK export ###
Q = pd.read_csv("FileList.csv", sep=",")
#####################################################

### Copy all data from original CSV to new XLSX doc ################
workbook = Workbook()
sheet = workbook.active

with pd.ExcelWriter("FileList_Temp.xlsx", mode="w") as writer:
    Q.to_excel(writer, sheet_name = "Sheet1", index = False)
#####################################################################

### Open the new xlsx formatted document and select sheet ###
NewExcel = load_workbook(filename="FileList_Temp.xlsx")
sheet = NewExcel.active
#############################################################

### Define how many rows/items there are ###
LoopFull = len(sheet["B"])
############################################

### Loop through all item numbers in coloumn B and change TEXT to HYPERLINK of text ############################################
for B_loop in range(LoopFull):
    if B_loop > 0:
        CellData = "B"+str(B_loop+1)
        CellExt = "D"+str(B_loop+1)
        
        Source = str(sheet[CellData].internal_value)
        SourceEXT = sheet[CellExt].internal_value
        
        DestinationExt = Extension(SourceEXT)
        Destination = "=hyperlink(%sitems/%s%s, %s%s%s)" % (chr(34), Source+DestinationExt, chr(34), chr(34), Source, chr(34))

        sheet[CellData] = Destination
################################################################################################################################


### Create a new Named Style for Excel ########################################################
if 'blue_clean' not in NewExcel.named_styles:
    blue_clean = NamedStyle(name="blue_clean")
    blue_clean.font = Font(color='000000', bold=True)
    blue_clean.fill = PatternFill(start_color='4adfff', end_color='4adfff', fill_type='solid')
    NewExcel.add_named_style(blue_clean)


# add style #
for cell in sheet["1:1"]:
    cell.style = 'blue_clean'
###############################################################################################


### Save the new interactive excel to correct location ###
NewExcel.save("InteractiveExcel\\Interactive_Excel_Prototype.xlsx")
##########################################################

### Clean up temp file ##############
os.remove("FileList_Temp.xlsx")
#####################################

###############################################################

